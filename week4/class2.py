def test3():
    import openpyxl
    # 엑셀 만들기
    wb = openpyxl.Workbook()
    # sheet 만들기
    ws = wb.create_sheet("회원정보")
    
    # 데이터 저장하기
    ws['A1'] = '이름'
    ws['B1'] = '나이'
    # ws['A2'] = '홍길동'
    # ws['B2'] = '01083088131'
    # ws['A3'] = "홍길순"
    # ws['B3'] = "01031028132"
    
    for i in range(1,6):
        ws[f'A{i+1}']=f'홍길동{i}'
        ws[f'B{i+1}']=20+i
    
    # 저장하기
    wb.save("member.xlsx")
    
# test3()



def test4():
    import openpyxl
    
    wb = openpyxl.load_workbook("member.xlsx")
    ws = wb['회원정보']
    
    list_item = []
    item={}
    
    header = [ws['A1'].value, ws['B1'].value]
    
    # print(header)
    
    for i in range(2,7):
        # print(ws[f'A{i}'].value , ws[f'B{i}'].value)
        item= {ws[f'A{i}'].value : ws[f'B{i}'].value}
        list_item.append(item)
    print(list_item)
    
    # for i in range(2,7):
    #     dic={}
    #     dic[header[0]] = ws[f'A{i}'].value
    #     dic[header[1]] = ws[f'B{i}'].value
    
# test4()


# =============================================================================
# Quiz) n명의 이름, 국어, 영어, 수학 점수를 입력받아
#       JSON 으로 저장 후 읽어서 개인 평균을 출력하기
# =============================================================================

def Quiz():
    import json
    #split()는 디폴트 값이 스페이스이다.
    item_list = []
    while(True):
        value = input("이름 국어 영어 수학점수를 입력하세요 [종료 : quit] : ")
        if value.upper() == 'QUIT':
            break
        val_list = value.split()
        dic = {}
        dic['name'] = val_list[0]
        dic['kor'] = val_list[1]
        dic['eng'] = val_list[2]
        dic['math'] = val_list[3]
        
        item_list.append(dic)
    
    with open("score.json","w", encoding="utf8") as f:
        json.dump(item_list, f, ensure_ascii=False)
    
    item_list = []
    with open("score.json","r",encoding="utf8") as f:
        item_list = json.load(f)
        
    # print(item_list)
    
    for item in item_list:
        sum = 0
        sum +=int(item['kor'])
        sum +=int(item['eng'])
        sum +=int(item['math'])
        print(item['name'], ":", sum/3)
    
Quiz()
